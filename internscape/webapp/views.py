from django.http import HttpResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.models import auth
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from .forms import LoginForm, CandidateRegisterForm, CompanyRegisterForm, JobForm, ApplicationForm, BlogForm
from .models import Profile, Job, Application, Blog
from django.contrib import messages
import openpyxl
from openpyxl.utils import get_column_letter

# Superadmin credentials
SUPERADMIN_EMAIL = 'superadmin'
SUPERADMIN_PASSWORD = '1234567@'

# Home page
def home(request):
    return render(request, 'webapp/index.html')

# Register a user
def candidate_register(request):
    form = CandidateRegisterForm()
    if request.method == "POST":
        form = CandidateRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            Profile.objects.create(user=user, user_type='candidate')
            messages.success(request, "Candidate Account Created Successfully")
            return redirect('my-login')
    context = {'form': form}
    return render(request, 'webapp/register.html', context)

def company_register(request):
    form = CompanyRegisterForm()
    if request.method == "POST":
        form = CompanyRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            Profile.objects.create(user=user, user_type='company')
            messages.success(request, "Company Account Created Successfully")
            return redirect('company-login')
    context = {'form': form}
    return render(request, 'compapp/compregister.html', context)

# Login a user
def candidate_login(request):
    form = LoginForm()
    if request.method == "POST":
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = request.POST.get('username')
            password = request.POST.get('password')
            # Check for superadmin credentials
            if username == SUPERADMIN_EMAIL and password == SUPERADMIN_PASSWORD:
                user = authenticate(request, username=username, password=password)
                if user is not None:
                    login(request, user)
                    messages.success(request, "Superadmin Login Successful")
                    return redirect('superadmin-dashboard')
            # Regular candidate login
            user = authenticate(request, username=username, password=password)
            if user is not None and user.profile.user_type == 'candidate':
                login(request, user)
                messages.success(request, "Login Successful")
                return redirect('dashboard')
            else:
                messages.error(request, "Invalid credentials or user type")
    context = {'form': form}
    return render(request, 'webapp/my-login.html', context)

def company_login(request):
    form = LoginForm()
    if request.method == "POST":
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = request.POST.get('username')
            password = request.POST.get('password')
            # Check for superadmin credentials
            if username == SUPERADMIN_EMAIL and password == SUPERADMIN_PASSWORD:
                user = authenticate(request, username=username, password=password)
                if user is not None:
                    login(request, user)
                    messages.success(request, "Superadmin Login Successful")
                    return redirect('superadmin-dashboard')
            # Regular company login
            user = authenticate(request, username=username, password=password)
            if user is not None and user.profile.user_type == 'company':
                login(request, user)
                messages.success(request, "Login Successful")
                return redirect('company-dashboard')
            else:
                messages.error(request, "Invalid credentials or user type")
    context = {'form': form}
    return render(request, 'compapp/complogin.html', context)

# Dashboards
@login_required(login_url='my-login')
def dashboard(request):
    return render(request, 'webapp/dashboard.html')

@login_required(login_url='company-login')
def company_dashboard(request):
    return render(request, 'compapp/compdashboard.html')

@login_required(login_url='my-login')
def superadmin_dashboard(request):
    pending_blogs = Blog.objects.filter(status='pending')
    approved_blogs = Blog.objects.filter(status='approved')
    context = {
        'pending_blogs': pending_blogs,
        'approved_blogs': approved_blogs,
    }
    return render(request, 'webapp/superadmin_dashboard.html', context)

# User logout
def user_logout(request):
    auth.logout(request)
    messages.success(request, "Logout Successful")
    return redirect("")

# Job-related views
@login_required(login_url='company-login')
def post_job(request):
    if request.method == "POST":
        form = JobForm(request.POST)
        if form.is_valid():
            job = form.save(commit=False)
            job.company = request.user
            job.save()
            return redirect('company-dashboard')
    else:
        form = JobForm()
    return render(request, 'compapp/post_job.html', {'form': form})

@login_required(login_url='my-login')
def view_jobs(request):
    jobs = Job.objects.all()
    return render(request, 'webapp/view_jobs.html', {'jobs': jobs})

@login_required(login_url='my-login')
def apply_for_job(request, job_id):
    job = get_object_or_404(Job, id=job_id)
    if request.method == "POST":
        form = ApplicationForm(request.POST, request.FILES)
        if form.is_valid():
            application = form.save(commit=False)
            application.candidate = request.user
            application.job = job
            application.save()
            return redirect('dashboard')
    else:
        form = ApplicationForm()
    return render(request, 'webapp/apply_for_job.html', {'form': form, 'job': job})

@login_required(login_url='company-login')
def job_postings_list(request):
    jobs = Job.objects.filter(company=request.user).order_by('-posted_at')
    context = {
        'jobs': jobs,
    }
    return render(request, 'compapp/job_postings_list.html', context)

@login_required(login_url='company-login')
def job_applications_detail(request, job_id):
    job = get_object_or_404(Job, pk=job_id)
    context = {
        'job': job,
    }
    return render(request, 'compapp/job_applications_detail.html', context)

@login_required(login_url='company-login')
def export_applications(request, job_id):
    job = get_object_or_404(Job, id=job_id)
    applications = job.applications.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Applications for {job.title}"
    headers = ['Candidate Name', 'Phone Number', 'Email', 'Cover Letter', 'Applied On']
    for col_num, header in enumerate(headers, 1):
        ws[f'{get_column_letter(col_num)}1'] = header
        ws[f'{get_column_letter(col_num)}1'].font = openpyxl.styles.Font(bold=True)
    for row_num, app in enumerate(applications, 2):
        ws[f'A{row_num}'] = app.name
        ws[f'B{row_num}'] = app.phone_number
        ws[f'C{row_num}'] = app.email
        ws[f'D{row_num}'] = app.cover_letter_file.url if app.cover_letter_file else 'N/A'
        ws[f'E{row_num}'] = app.applied_at.strftime('%Y-%m-%d %H:%M')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=applications_{job.id}.xlsx'
    wb.save(response)
    return response

# Blog-related views
@login_required(login_url='my-login')
def view_blogs(request):
    blogs = Blog.objects.filter(status='approved').order_by('-created_at')
    return render(request, 'webapp/blogs.html', {'blogs': blogs})

@login_required(login_url='my-login')
def create_blog(request):
    if request.method == "POST":
        form = BlogForm(request.POST)
        if form.is_valid():
            blog = form.save(commit=False)
            blog.author = request.user
            blog.status = 'pending'
            blog.save()
            messages.success(request, "Blog submitted for approval")
            return redirect('view-blogs')
    else:
        form = BlogForm()
    return render(request, 'webapp/create_blog.html', {'form': form})


def approve_blog(request, blog_id):
    blog = get_object_or_404(Blog, id=blog_id)
    blog.status = 'approved'
    blog.save()
    messages.success(request, "Blog approved")
    return redirect('superadmin-dashboard')


def reject_blog(request, blog_id):
    blog = get_object_or_404(Blog, id=blog_id)
    blog.status = 'rejected'
    blog.save()
    messages.success(request, "Blog rejected")
    return redirect('superadmin-dashboard')


def delete_blog(request, blog_id):
    blog = get_object_or_404(Blog, id=blog_id)
    blog.delete()
    messages.success(request, "Blog deleted")
    return redirect('superadmin-dashboard')